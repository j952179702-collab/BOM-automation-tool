# 111.py - 仪表自动化处理器（修改版）

import os
import sys
import re
import logging
from pathlib import Path
from typing import Optional, Tuple
from venv import logger
import numpy as np
from numpy.random import f
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from base_processor import BaseProcessor
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import sys

# ====================== QTextEditLogger ======================
class QTextEditLogger(logging.Handler):
    def __init__(self):
        super().__init__()
        self.widget = None

    def set_widget(self, widget):
        self.widget = widget

    def emit(self, record):
        msg = self.format(record)
        if self.widget:
            self.widget.append(msg)

# ====================== 主功能类 ======================
class InstrumentAutomationProcessor(BaseProcessor):
    def __init__(self, input_file):
        super().__init__(input_file)
        self.input_file = input_file
        self.df = None
        self.df_sort = None
        self.processed_workbook = None
        self.logger = None
        self.log_handler = QTextEditLogger()
        # 注意：不要在这里调用 load_csv，因为它需要 logger 已经初始化
    
    def save_processed_file(self, save_path: str) -> bool:
        """实现抽象方法：保存处理结果"""
        if self.df_sort is None:
            self.logger.error("❌ 没有可保存的数据（df_sort 为 None）")
            return False

        try:
            # 使用 pandas 保存为 Excel
            self.df_sort.to_excel(save_path, index=False, engine='openpyxl')
            self.logger.info(f"✅ 结果已保存到: {save_path}")
            return True
        except Exception as e:
            self.logger.error(f"❌ 保存文件失败: {str(e)}")
            return False

    def setup_logging(self, text_edit_widget):
        """由 UI 传入 QTextEdit，初始化日志"""
        self.logger = logging.getLogger(f'Processor_{id(self)}')
        self.logger.setLevel(logging.INFO)
        self.logger.handlers.clear()

        # 设置 handler
        self.log_handler.set_widget(text_edit_widget)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S')
        self.log_handler.setFormatter(formatter)
        self.logger.addHandler(self.log_handler)

        # 控制台输出
        console = logging.StreamHandler()
        console.setFormatter(formatter)
        self.logger.addHandler(console)

        # ✅ 现在才能用 logger
        self.logger.info("日志系统已启动")

    def load_csv(self) -> pd.DataFrame:
        try:
            encodings = ['utf-8', 'gbk']

            for enc in encodings:
                try:
                    df = pd.read_csv(self.input_file, encoding=enc)
                    print(f"✅ 成功使用编码: {enc}")
                    break  # 成功就读取并退出循环
                except UnicodeDecodeError as e:
                    print(f"❌ {enc} 解码失败: {e}")
                    continue
            else:
                # 所有编码都失败
                raise ValueError("❌ 所有编码都无法解析该文件，请检查文件格式或编码")
            self.logger.info(f"✅ 成功加载CSV文件: {self.input_file}")
            return df
        except Exception as e:
            self.logger.error(f"❌ 文件加载失败: {str(e)}")
            raise

    def resource_path(self, relative_path):
        """ 获取资源的绝对路径，支持开发环境和 PyInstaller 打包 """
        try:
            # PyInstaller 打包后的临时路径
            base_path = sys._MEIPASS
        except AttributeError:
            # 开发环境下：使用当前脚本所在目录的父目录（即项目根目录）
            # 假设该方法定义在 ui/ 目录下的某个类中
            base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)
        
    def extract_chinese(self,text) -> str:
        # 提取所有中文字符
        if pd.isna(text) or not isinstance(text, str):
            text = str(text) if not pd.isna(text) else ""
        chinese = re.findall(r"[\u4e00-\u9fa5]+", text)
        return "".join(chinese) if chinese else "未知"  

    def generate_code(self):
        try:
            if self.df is None:
                self.df = self.load_csv()
            self.df_sort = self.df.sort_values(by=self.df.columns[1], ascending=False).copy()
            self.logger.info(f"已排序，共 {len(self.df_sort)} 行数据")

            # 初始化列（防止 KeyError）
            for col in ["名称", "量程", "表壳材质", "仪表材质", "探杆长度", "安装方式", "仪表类型","线缆长度"]:
                if col not in self.df_sort.columns:
                    self.df_sort[col] = "编码错误"

            # 量程映射
            range_mapping_rules = {
                "双金属温度计": {"-40~0": "L", "0~100": "M", "0~200": "H"},
                "压力变送器": {"-0.1~0.1": "A", "0~0.25": "B", "0~0.5": "C", "0~1.0": "D"},
                "温度变送器": {"-40~0": "L", "0~100": "M", "0~200": "H"},
                "压力表": {"-0.1~0.1": "L", "0~1.0": "M", "0~0.5": "N", "0~0.25": "O", "0~1.6": "G"}
            }

            flat_map = {}
            for inst_type, ranges in range_mapping_rules.items():
                for range_val, code in ranges.items():
                    flat_map[(inst_type, range_val)] = code

            range_mapped = self.df_sort["量程"].astype(str).map(flat_map)
            range_val_series = pd.Series(
                np.where(
                    range_mapped.isna(),
                    self.df_sort["量程"].astype(str),  # 回退到原始量程(Y)
                    range_mapped                     # 使用映射编码
                ),
                index=self.df_sort.index
            ).replace("nan", "错误")  # 处理可能的 'nan' 字符串
            range_val = pd.Series(range_val, index=self.df_sort.index).replace("nan", "错误")
            
            # 材质映射
            material_mapping = {
                "304": "SS", "316L": "SS1", "TA2": "Ti", "Ta": "Ta",
                "HC": "HC", "2205": "SS2", "压铸铝": "C", "编码错误": "ERR","UPVC":"UPVC","TAN":"Ta"
            }
            self.shell_material = self.df_sort["表壳材质"].fillna("编码错误").map(material_mapping).fillna("错误")
            self.material_value = self.df_sort["仪表材质"].fillna("编码错误").map(material_mapping).fillna("错误")

            # 其他字段
            probe = self.df_sort["探杆长度"].fillna("0").astype(int) if "探杆长度" in self.df_sort.columns else " "
            install_type = self.df_sort["安装方式"].fillna("编码错误").astype(str) if "安装方式" in self.df_sort.columns else " "
            long = self.df_sort["线缆长度"].fillna("编码错误").astype(str) if "线缆长度" in self.df_sort.columns else " "

            def get_param_text(medium):
                if medium in medium_map:
                    # 按中文分号拆分，去掉空，再用 \n 连接
                    parts = [p.strip() for p in medium_map[medium].split("；") if p.strip()] 
                    return "\n".join(parts)  # ✅ 用换行符连接，存入一个格子
                else:
                    return "无"
            parameter_ins = {
                "压力表": "（*）. 精度：1.0%FS；\n（*）. 防护等级：IP65；\n（*）. 安装方式：径向直接式",
                "压力变送器": "（*）.精度：0.5%FS；\n（*）.输出信号：4-20mA；\n（*）.电压：两线制；\n（*）.通讯协议：无；\n（*）.现场显示：一体式多功能LCD显示表；\n（*）.防护等级：IP65；\n（*）.防爆等级：无；\n（*）.电气密封接口：M20*1.5；\n（*）.安装方式：螺纹；",
                "法兰液位仪表": "（*）. 精度：0.5%FS；\n（*）. 输出信号：4-20mA，二线制；\n（*）. 通讯协议：无；\n（*）. 防护等级：IP65；\n（*）. 电气密封接口：M20*1.5；\n（*）. 防爆等级：无",
                "投入液位仪表": "（*）. 精度：0.5%FS；\n（*）. 输出信号：4-20mA，二线制；\n（*）. 通讯协议：无；\n（*）. 现场显示：LCD显示，带调零功能；\n（*）. 防护等级：IP65；\n（*）. 电气密封接口：M20*1.5；\n（*）. 防爆等级：无",
                "液位计开关": "（*）.输出信号：开关量；\n（*）.防护等级：IP65；\n（*）.防爆等级：/；\n（*）.电气密封接口：不带接线盒；\n（*）.安装/接管方式：浸没",
                "电磁流量计":"（*）.精度：0.5%FS；\n（*）.输出信号：DC24V，4-20mA；\n（*）.电压：四线制；\n（*）.通讯协议：modbus485；\n（*）.现场显示：一体式多功能LCD显示表；\n（*）.防护等级：IP65；\n（*）.防爆等级：无；\n（*）.电气密封接口：M20*1.5；",
                "涡街流量计":"（*）.精度：1.0级；\n（*）.输出信号：DC24V，4-20mA；\n（*）.电压：四线制；\n（*）.通讯协议：modbus485；\n（*）.现场显示：一体式多功能LCD显示表；\n（*）.防护等级：IP65；\n（*）.防爆等级：无；\n（*）.电气密封接口：M20*1.5；",
                "热式流量计":"（*）.精度：1.0级；\n（*）.输出信号：DC24V，4-20mA；\n（*）.电压：四线制；\n（*）.通讯协议：modbus485；\n（*）.现场显示：一体式多功能LCD显示表；\n（*）.防护等级：IP65；\n（*）.防爆等级：无；\n（*）.电气密封接口：M20*1.5；",
                "温度变送器":"（*）.精度：0.5%FS；\n（*）.输出信号：4-20mA；\n（*）.电压：DC24V,两线制；\n（*）.通讯协议：/；\n（*）.现场显示：无；\n（*）.防护等级：IP65；\n（*）.防爆等级：/；\n（*）.电气密封接口：M20*1.5；\n（*）.保护管类型：螺纹式直型保护管；\n（*）.保护管直径：φ12；\n（*）.备注：配聚四氟乙烯垫片1个；\n（*）.安装/接管方式：固定G1/2外螺纹,PN10；",
                "双金属温度计":"（*）.精度：1.0%FS；\n（*）.现场显示：表盘直径φ100，表头填充硅油；\n（*）.防护等级：IP65；\n（*）.形式：耐震，万向型\n（*）.安装方式：轴向安装\n（*）.鞘直径：6mm；\n（*）.鞘安装接头：G1/2外螺纹，PN10；\n（*）.保护管类型：螺纹式直型保护管；\n（*）.保护管直径：φ12；\n（*）.保护管与鞘连接：G1/2内螺纹，PN10；\n（*）.安装/接管方式：固定G1/2外螺纹；",
                "分析仪表":"（*）.输出信号：4-20mA；\n（*）.电压：DC24V，两线制；\n（*）.通讯协议：RS485；\n（*）.现场显示：一体式多功能LCD显示表；\n（*）.防护等级：IP65；\n（*）.防爆等级：无；\n（*）.电气密封接口：2-M20*1.5；",
                "法兰浮子流量计":""
                }

            self.df_sort['参数2'] = self.df_sort['名称'].apply(lambda x: self.extract_chinese(x)).map(parameter_ins)
            unmatched_mask = self.df_sort['参数2'].isna()
            if unmatched_mask.any():
                unmatched = self.df_sort.loc[unmatched_mask, '名称'].unique()
                self.logger.error(f"⚠️ 以下仪表名称未在标准参数库中找到：{unmatched.tolist()}")
                self.df_sort['参数2'] = self.df_sort['参数2'].fillna("")
            else:
                self.logger.info("✅ 所有仪表名称均已匹配标准参数")

            # 生成 SKU
            sku_list = []
            mediumn_path = self.resource_path("dataset\\medium.xlsx")
            medium_df =  pd.read_excel(mediumn_path, engine='openpyxl')
            medium_df.dropna(subset=['介质', '参数'], inplace=True)
            medium_map = dict(zip(medium_df["介质"], medium_df["参数"]))
            self.df_sort["参数1"] = self.df_sort["介质"].apply(get_param_text) if "介质" in self.df_sort.columns else "无"    
            for idx in self.df_sort.index:
                param1_val = str(self.df_sort.loc[idx, '参数1'])
                param2_val = str(self.df_sort.loc[idx, '参数2'])
                self.ins_name =  self.extract_chinese(str(self.df_sort.loc[idx, "名称"])).strip()
                range_val_str = str(self.df_sort.loc[idx, "量程"])
                key = (self.ins_name,range_val_str)
                if key in flat_map:
                    self.r_val = flat_map[key]
                    self.logger.info(f"✅ 匹配成功: {key} → {self.r_val}")
                else:
                    self.r_val = range_val_str  # 回退到原始量程
                    self.logger.error(f"❌ 未匹配: {key} → 使用原始值")
                mat_val = self.material_value.loc[idx]
                self.shell_mat = self.shell_material.loc[idx]
                self.prob = probe.loc[idx]
                self.install_val = install_type.loc[idx]
                self.long_val = long.loc[idx]

                shell_value = self.df_sort.loc[idx,'表壳材质']
                material_value = self.df_sort.loc[idx,'仪表材质']
                
                try:
                    if self.ins_name == "双金属温度计":
                        sku = f"WSS-AOA{self.r_val}{self.shell_mat}-12{mat_val}{self.prob}-LWG1/2"
                        self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：{shell_value}；\n（*）.保护管材质：{material_value}（含接头）"
                        self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{range_val_str}℃；\n（*）.保护管长度：{self.prob}mm；\n{param2_val}"
                    elif self.ins_name == "温度变送器":
                        sku = f"SWBZ-A0-24D2NNA0P{self.r_val}0-12{mat_val}{self.prob}-LWG1/2"
                        self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.保护管材质：{material_value}（含接头）"
                        self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{range_val_str}℃；\n（*）.保护管长度：{self.prob}mm；\n{param2_val}"
                    elif self.ins_name == "压力表":
                        inst_type = self.df_sort.loc[idx, "仪表类型"]
                        if str(inst_type) == "耐震":
                            sku = f"YTHN-A1A{self.r_val}{self.shell_mat}-{mat_val}-LWG1/2W"
                            self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：{shell_value}；\n（*）.接液材质：{material_value}"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{range_val_str}MPa（G）；\n（*）.形式：普通型耐震；\n（*）.现场显示：表盘直径φ100，表头填充 silicone；\n{param2_val}\n（*）.安装/接管方式：固定G1/2外螺纹,PN10；"
                        elif str(inst_type) == "耐震隔膜":
                            if mat_val == "PP":
                                sku = f"YMN-A1A{self.r_val}PP-PP-LWG1/2N"
                                self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：{shell_value}；\n（*）.膜片材质：{material_value}"
                                self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{range_val_str}MPa（G）；\n（*）.现场显示：表盘直径φ60，表头填充 silicone；\n（*）.形式：耐震隔膜型；\n{param2_val}\n（*）.安装/接管方式：固定G1/2内螺纹,PN10；"

                            else:
                                    sku = f"YMN-A1A{self.r_val}{self.shell_mat}-{mat_val}-FL120RF1.0"
                                    self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：{shell_value}；\n（*）.膜片材质：{material_value}"
                                    self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{range_val_str}MPa（G）；\n（*）.形式：耐震隔膜型；\n（*）.现场显示：表盘直径φ60，表头填充 silicone；\n{param2_val}\n（*）.安装/接管方式：法兰连接，DN20,PN10"
                    elif self.ins_name == "压力变送器":
                        if self.r_val in ["0~0.25", "B"]:
                            sku = f"PT-H-B-1-{mat_val}-50-NSR"
                            self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{range_val_str}MPa（G）；\n{param2_val}\n（*）.安装/接管方式：法兰DN50，PN10 HG/T-20592-2009，RF，B型；"
                        else:
                            sku = f"PT-H-{self.r_val}-1-{mat_val}-G1/2-NSR"
                            self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{range_val_str}MPa（G）；\n{param2_val}\n（*）.安装/接管方式：固定G1/2外螺纹,PN10；"
                    elif self.ins_name == "电磁流量计":
                        diameter_dianci = {
                            "DN15":"0.32~1.27",
                            "DN20":"0.57~2.26",
                            "DN25":"0.88~3.53",
                            "DN32":"1.45~5.79",
                            "DN40":"2.3~9.0",
                            "DN50":"3.5~14.1",
                            "DN65":"6.0~23.9",
                            "DN80":"9~36.2",
                            "DN100":"14.1~56.5",
                            "DN125":"22.1~88.4",
                            "DN150":"31.8~127",
                            "DN200":"56.5~226"
                        }
                        diameter1 = diameter_dianci.get(self.install_val, "未知")
                        sku = f"EMF-1C-0-A{mat_val}CS-FL{self.install_val.split('DN')[-1]}RF1.0"
                        self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.电极材质：{material_value}；\n（*）.接液材质：PTFE；\n（*）.法兰材质：碳钢"
                        self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.介质流量：{diameter1}m³/h；\n{param2_val}\n（*）.安装/接管方式：法兰DN{self.install_val.split('DN')[-1]}，PN10 HG/T-20592-2009，RF，B型"
                    elif self.ins_name == "热式流量计":
                        diameter_reshi = {
                            "DN15":"0.65~65",
                            "DN25":"1.75~175",
                            "DN32":"2.9~290",
                            "DN40":"4.5~450",
                            "DN50":"7~700",
                            "DN65":"12~1200",
                            "DN80":"18~1800",
                            "DN100":"28~2800",
                            "DN125":"44~4400",
                            "DN150":"63~6300",
                            "DN200":"100~1000"
                        }
                        diameter2 = diameter_reshi.get(self.install_val, "未知")
                        if int(self.install_val.split('DN')[-1]) <= 80:
                            sku = f"TFC-A{diameter2.split('~')[-1]}-FL{self.install_val.split('DN')[-1]}"
                            self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}；\n（*）.法兰材质：304"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.介质流量：{diameter2}m³/h；\n{param2_val}\n（*）.安装/接管方式：法兰DN{self.install_val.split('DN')[-1]}，PN10 HG/T-20592-2009，RF，B型；"
                        else:
                            sku = f"TFC-A{diameter2.split('~')[-1]}-LWG1/2-{self.install_val.split('DN')[-1]}"
                            self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}；\n（*）.螺纹材质：304"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.介质流量：{diameter2}m³/h；\n{param2_val}\n（*）.安装/接管方式：法兰DN{self.install_val.split('DN')[-1]}，PN10 HG/T-20592-2009，RF，B型；"
                    elif self.ins_name == "法兰浮子流量计":
                        if "DN" in self.install_val:
                            # 提取 DN 后的数字，比如 DN50 -> 50
                            try:
                                dn_size = self.install_val.split('DN')[-1].strip()
                                # 进一步提取数字（防止有单位）
                                dn_number = re.search(r'\d+', dn_size)
                                dn_display = dn_number.group() if dn_number else dn_size
                            except:
                                dn_display = "XX"

                            sku = f"FF-{mat_val}-{self.r_val.replace('-','~').split('~')[-1]}-FL{dn_display}"
                            self.df_sort.loc[idx, '材质'] = f"（*）.材质：{material_value}"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.介质流量：{self.r_val}m³/h；\n（*）.法兰DN{dn_display}，PN10 HG/T-20592-2009，RF，B型；"
                        else:
                            # 如果没有 DN，说明是承插焊
                            # 你可以选择提取数字，或直接用文本
                            try:
                                size_match = re.search(r'\d+', self.install_val)
                                size_display = size_match.group() if size_match else "XX"
                            except:
                                size_display = "XX"

                            sku = f"FF-{mat_val}-{self.r_val.replace('-','~').split('~')[-1]}-CC{size_display}"
                            self.df_sort.loc[idx, '材质'] = f"（*）.材质：{material_value}"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.介质流量：{self.r_val}m³/h；\n{param2_val}\n（*）.安装/接管方式：承插{size_display}；"
                    elif self.ins_name =="涡街流量计":
                        diamter_wojie = {
                            "DN15":"4~28",
                            "DN20":"6~40",
                            "DN25":"8~50",
                            "DN32":"13~130",
                            "DN40":"25~180",
                            "DN50":"35~300",
                            "DN65":"50~500",
                            "DN80":"80~800",
                            "DN100":"120~1200",
                            "DN125":"180~1800",
                            "DN150":"320~2800",
                            "DN200":"560~6000"
                        }
                        sku =  f"VF-1C-{mat_val}{mat_val}-FL{self.install_val.split('DN')[-1]}RF1.0"
                        diameter3 = diamter_wojie.get(self.install_val, "未知")
                        self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}；\n（*）.法兰材质：{material_value}"
                        self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.介质流量：{diameter3}m³/h；\n{param2_val}\n（*）.安装/接管方式：法兰DN{self.install_val.split('DN')[-1]}"
                    elif self.ins_name == "法兰液位仪表":
                        if (self.df_sort.loc[idx, "仪表类型"] == "磁翻板") or (self.df_sort.loc[idx, "仪表形式"] == "磁翻板"):
                            if "KP" in install_type:
                                sku = f"LG-FQC-{self.r_val.replace('-','~').split('~')[-1]}-{mat_val}{mat_val}-KP50"
                                self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}；\n（*）.法兰材质：{material_value}"
                                self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{self.r_val}mm；\n{param2_val}\n（*）.安装/接管方式：卡盘直径50.5；"
                            else:
                                sku = f"LG-FQC-{self.r_val.replace('-','~').split('~')[-1]}-{mat_val}{mat_val}-FL25RF1.0"
                                self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}；\n（*）.法兰材质：{material_value}"
                                self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{self.r_val}mm；\n{param2_val}\n（*）.安装/接管方式：卡盘直径50.5；"
                        elif (self.df_sort.loc[idx, "仪表类型"] == "单法兰") or (self.df_sort.loc[idx, "仪表形式"] == "单法兰"):
                            sku = f"LG-DFC-{self.r_val.replace('-','~').split('~')[-1]}-SS{mat_val}-FL50RF1.0"
                            self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}；\n（*）.法兰材质：304"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{self.r_val}mm；\n{param2_val}\n（*）.安装/接管方式：法兰DN50，PN10 HG/T-20592-2009，RF，B型；"
                        elif (self.df_sort.loc[idx, "仪表类型"] == "双法兰") or (self.df_sort.loc[idx, "仪表形式"] == "双法兰"):
                            sku = f"LG-SFC-{self.r_val.replace('-','~').split('~')[-1]}-SS{mat_val}-FL50RF1.0"
                            self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}；\n（*）.法兰材质：304"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{self.r_val}mm；\n（*）.现场显示：LCD显示，带调零功能；\n{param2_val}\n（*）.安装/接管方式：法兰DN50，PN10 HG/T-20592-2009，RF，B型；"
                    elif self.ins_name == "投入液位仪表":
                        sku = f"LG-TRC-{self.r_val.replace('-','~').split('~')[-1]}-{mat_val}-{mat_val}-TR"
                        self.df_sort.loc[idx,'材质'] = f"（*）.表壳材质：压铸铝；\n（*）.接液材质：{material_value}"
                        self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：{self.r_val}mm；\n{param2_val}\n（*）.安装/接管方式：投入式；"
                    elif self.ins_name == "液位计开关":
                        sku = f"CFBS-{self.long_val}-PP-0-0"
                        self.df_sort.loc[idx,'材质'] = f"（*）.接液材质：{material_value}"
                        self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.线缆长度{self.long_val}；\n{param2_val}\n（*）.安装/接管方式：投入式；"
                    elif self.ins_name == "分析仪表":
                        if self.df_sort.loc[idx,"仪表类型"] == "Ω":
                            if re.search(r'\b2000\b',self.r_val):
                                sku = f"EC-W1-A1/2000-SS1-LWG1/2"
                                self.df_sort.loc[idx,'材质'] = f"（*）.接液材质：{material_value}"
                                self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：1-2000us/cm，k=1；\n（*）.精度：0.5%FS；\n（*）.线长：{self.long_val}\n{param2_val}\n（*）.安装/接管方式：固定G1/2外螺纹；"
                            elif re.search(r'\b20000\b',self.r_val):
                                sku = f"EC-W1-A10/20000-SS1-LWG1/2"
                                self.df_sort.loc[idx,'材质'] = f"（*）.接液材质：{material_value}"
                                self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：10-20000us/cm，k=10；\n（*）.线长：{self.long_val}\n（*）.精度：0.5%FS；\n{param2_val}\n（*）.安装/接管方式：固定G1/2外螺纹；"
                            elif re.search(r'\b30000\b',self.r_val):
                                sku = f"EC-W1-A10/600000-SS1-LWG1/2"
                                self.df_sort.loc[idx,'材质'] = f"（*）.接液材质：聚砜"
                                self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：30000~600000us/cm，k=30；\n（*）.精度：0.5%FS；\n（*）.线长：{self.long_val}\n{param2_val}\n（*）.安装/接管方式：固定G3/4螺纹；"
                            else:
                                sku = "编码失败"
                        elif self.df_sort.loc[idx,"仪表类型"] == "PH":
                            sku = f"PH-0-DC24VA/14-GL-LWG3/4"
                            self.df_sort.loc[idx,'材质'] = f"（*）.接液材质：玻璃"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：0-14；\n（*）.精度：±0.01；\n（*）.线长：{self.long_val}\n{param2_val}\n（*）.安装/接管方式：带支架，G3/4螺纹；"
                        elif self.df_sort.loc[idx,"仪表类型"] == "DO":
                            sku = f"DO-0-A/020-SS1-JM"
                            self.df_sort.loc[idx,'材质'] = f"（*）.保护管材质：PP；\n（*）.接液材质：316L；"
                            self.df_sort.loc[idx,'参数'] = f"{param1_val}\n（*）.量程：0-20mg/L；\n（*）.精度：±0.5%FS；（*）.线长：{self.long_val}；\n{param2_val}\n（*）.安装/接管方式：浸没；"
                    else:
                        sku = f"无编码"
                        self.df_sort.loc[idx,"参数"] = f"\n无编码；"
                        self.df_sort.loc[idx,'材质'] = f"无编码"
                    sku_list.append(sku)

                except Exception as e:
                    sku = f"错误_{idx}"
                    sku_list.append(sku)
                    self.logger.error(f"行 {idx} 生成失败: {e}")
            self.df_sort['*SKU编号'] = sku_list
            self.df_sort['*申购单类型'] = '阀门仪表-仪表'
            self.logger.info("✅ SKU 生成完成！")
            return self.df_sort
        except Exception as e:
            self.logger.error(f"❌ 生成 SKU 时出错: {str(e)}")
            raise
    
    def set_metadata(self, applicant, date, project_number):
        """设置申购人、日期、项目号等元数据"""
        if self.df_sort is not None:
            self.df_sort['申购人'] = applicant
            self.df_sort['申购日期'] = date
            self.df_sort['项目号'] = project_number
            self.df_sort['所属项目'] = project_number
            self.df_sort['申购单类型'] = '阀门仪表-仪表'
        else:
            raise ValueError("df_sort 未初始化，请先加载数据")
    
    def get_note(self, df):
        """
        为 DataFrame 生成备注列：项目号 + 仪表名称
        """      
        required_columns = ['项目号', '仪表名称']  
        # ✅ 详细的列检查
        missing_cols = [col for col in required_columns if col not in df.columns]
        if missing_cols:
            error_msg = f"❌ 生成备注失败：缺少必要列 {missing_cols}"
            self.logger.error(error_msg)
            self.logger.info(f"📋 当前DataFrame的列名: {list(df.columns)}")
            raise ValueError(error_msg)

        try:
            # ✅ 安全转换为字符串，处理 NaN
            project_num = df['项目号'].fillna('').astype(str)
            ins_name = df['仪表名称'].fillna('').astype(str)
            weihao = df['仪表位号'].fillna('').astype(str)           
            # 确保字符串连接时有空格分隔
            df['备注'] = project_num.str.strip()+ ins_name.str.strip()+ weihao.str.strip()
            now = datetime.datetime.now().strftime("%Y%m%d")
            df['申购单备注'] = "【" + df['项目号'].astype(str) + "】仪表申购单-" + now
            
            # ✅ 检查备注列是否成功创建
            if '备注' not in df.columns:
                raise ValueError("备注列未成功创建")
            
            self.logger.info(f"✅ 备注列生成成功，前3个值: {df['备注'].head(3).tolist()}")
            self.logger.info(f"✅ 备注列数据类型: {df['备注'].dtype}")
            return df
        except Exception as e:
            error_msg = f"❌ 生成备注列时发生未知错误: {e}"
            self.logger.error(error_msg)
            raise
    

    def merge_by_SKU(self):
        self.df_sort['备注'] = self.df_sort['备注'].fillna('')
        aggregation = {
            '项目号': 'first',
            '仪表类型': 'first',
            '参数': 'first',
            '材质': 'first',
            '申购单备注': 'first',
            '备注': lambda x:'\n'.join(item for item in x if item.strip() != ''),
            '申购人': 'first',
            '申购日期': 'first',
            '*申购单类型': 'first'
        }
        cols_needed = list(aggregation.keys())
        df_clean = self.df_sort[['*SKU编号'] + cols_needed].copy()
        self.df_group = df_clean.groupby('*SKU编号', as_index=False).agg(aggregation)
        self.df_group['*申请数量'] = self.df_sort.groupby('*SKU编号').size().values
        self.df_group['所属项目'] = self.df_group['项目号']
        self.df_group = self.df_group.rename(columns={'项目号': '项目编号',
                                                      '申购日期': '需求日期',
                                                      '仪表类型':'产品名称'
                                                      })
        template_path = self.resource_path("dataset/申购单导入模板.xlsx")
        template_df =  pd.read_excel(template_path, engine='openpyxl',header = 1)
        template_columns = [str(col).strip().replace('\n', '').replace('\r', '') for col in template_df.columns]
        for col in template_columns:
            if col not in self.df_group.columns:
                self.df_group[col] = ''

        self.df_output = self.df_group[template_columns].reset_index(drop=True)  # ✅ 关键！
        mapping = {
            'V': '涡街流量计',
            'E': '电磁流量计',
            'F': '浮子流量计',
            'Ω':'电导率仪',
            'PH':'pH计',
            '耐震':'耐震压力表',
            '耐震隔膜':'耐震隔膜压力表',
            '磁翻板':'磁翻板液位计',
            '双法兰':'双法兰液位计',
            '单法兰':'单法兰液位计',
            'DO':'溶氧仪'
            }
        self.df_output['产品名称'] = self.df_output['产品名称'].replace(mapping) 
        template_path = self.resource_path("dataset/申购单导入模板.xlsx")
        book = load_workbook(template_path)
        sheet = book.active
        start_row = 3
        for row_idx,row in self.df_output.iterrows():
            for c_idx,value in enumerate(row,1):
                cell = sheet.cell(row = start_row + row_idx, column = c_idx,value = value)
                cell.alignment = Alignment(
                    wrap_text=True,           # 自动换行
                    vertical='center',        # 垂直居中
                    horizontal='center'       # 水平居中（可选）
                )
                if c_idx == sheet.max_column:  # 如果是最后一列，设置自动换行
                    cell.alignment = Alignment(wrap_text=True)
        self.logger.info(f"✅ 已生成申购单")
        self.book = book
        return book  



    def process(self, applicant, date, project_number):
        try:
            # 1. 生成SKU编码 
            self.df_sort = self.generate_code()
            if self.df_sort is None:
                raise ValueError("SKU编码生成失败，df_sort 为空")
            self.logger.info(f"✅ 生成SKU编码完成，数据形状: {self.df_sort.shape}")
            
            # 2. 设置元数据
            self.df_sort['申购人'] = applicant
            self.df_sort['申购日期'] = date
            self.df_sort['*申购单类型'] = '阀门仪表-仪表'
            # 确保项目号列存在并设置值
            if '项目号' not in self.df_sort.columns:
                self.df_sort['项目号'] = ''
            self.df_sort['项目号'] = project_number
            self.logger.info(f"✅ 设置元数据完成，项目号: {project_number}")
            
            # 3. 生成备注列
            if '仪表名称' not in self.df_sort.columns:
                # 假设仪表名称就是从“名称”列里提取的中文部分
                self.df_sort['仪表名称'] = self.df_sort['名称'].apply(lambda x: self.extract_chinese(str(x)))
            self.logger.info(f"开始生成备注列，当前列: {self.df_sort.columns.tolist()}")
            self.df_sort = self.get_note(df=self.df_sort)
            if '备注' not in self.df_sort.columns:
                raise ValueError("备注列生成失败")
            self.merge_by_SKU()
            
            self.logger.info("✅ 文件处理完成")
            return True
        except Exception as e:
            self.logger.error(f"❌ 处理失败: {str(e)}")
            return False