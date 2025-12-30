import os
import sys
import re
import logging
from pathlib import Path
from typing import Optional, Tuple
from venv import logger
import numpy as np
from numpy.random import f
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from base_processor import BaseProcessor
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys
from ValveMatch import ValveProtocolMatcher,ParameterFiller

# ====================== QTextEditLogger ======================
class QTextEditLogger(logging.Handler):
    def __init__(self):
        super().__init__()
        self.widget = None

    def set_widget(self, widget):
        self.widget = widget

    def emit(self, record):
        msg = self.format(record)
        if self.widget:
            self.widget.append(msg)

# ====================== 主功能类 ======================
class ValveAutomationProcess(BaseProcessor):
    def __init__(self, input_file):
        super().__init__(input_file)
        self.input_file = input_file
        self.df = None
        self.df_sort = None
        self.processed_workbook = None
        self.logger = None
        self.log_handler = QTextEditLogger()
    def save_processed_file(self, save_path: str) -> bool:
        """实现抽象方法：保存处理结果"""
        if self.df_sort is None:
            self.logger.error("❌ 没有可保存的数据（df_sort 为 None）")
            return False

        try:
            # 使用 pandas 保存为 Excel
            self.df_sort.to_excel(save_path, index=False, engine='openpyxl')
            self.logger.info(f"✅ 结果已保存到: {save_path}")
            return True
        except Exception as e:
            self.logger.error(f"❌ 保存文件失败: {str(e)}")
            return False

    def setup_logging(self, text_edit_widget):
        """由 UI 传入 QTextEdit，初始化日志"""
        self.logger = logging.getLogger(f'Processor_{id(self)}')
        self.logger.setLevel(logging.INFO)
        self.logger.handlers.clear()

        # 设置 handler
        self.log_handler.set_widget(text_edit_widget)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S')
        self.log_handler.setFormatter(formatter)
        self.logger.addHandler(self.log_handler)

        # 控制台输出
        console = logging.StreamHandler()
        console.setFormatter(formatter)
        self.logger.addHandler(console)

        # ✅ 现在才能用 logger
        self.logger.info("日志系统已启动")

    def load_csv(self) -> pd.DataFrame:
        try:
            encodings = ['utf-8', 'gbk']

            for enc in encodings:
                try:
                    df = pd.read_csv(self.input_file, encoding=enc)
                    print(f"✅ 成功使用编码: {enc}")
                    break  # 成功就读取并退出循环
                except UnicodeDecodeError as e:
                    print(f"❌ {enc} 解码失败: {e}")
                    continue
            else:
                # 所有编码都失败
                raise ValueError("❌ 所有编码都无法解析该文件，请检查文件格式或编码")
            self.logger.info(f"✅ 成功加载CSV文件: {self.input_file}")
            return df
        except Exception as e:
            self.logger.error(f"❌ 文件加载失败: {str(e)}")
            raise

    def resource_path(self, relative_path):
        """ 获取资源的绝对路径，支持开发环境和 PyInstaller 打包 """
        try:
            # PyInstaller 打包后的临时路径
            base_path = sys._MEIPASS
        except AttributeError:
            # 开发环境下：使用当前脚本所在目录的父目录（即项目根目录）
            # 假设该方法定义在 ui/ 目录下的某个类中
            base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)
        
    def extract_chinese(self,text) -> str:
        # 提取所有中文字符
        if pd.isna(text) or not isinstance(text, str):
            text = str(text) if not pd.isna(text) else ""
        chinese = re.findall(r"[\u4e00-\u9fa5]+", text)
        return "".join(chinese) if chinese else "未知"  
    
    def generate_code(self):
        if self.df is None:
            self.df = self.load_csv()
        
        # 1. 初始排序（按第2列降序）
        # 建议：如果知道第2列的具体名称，最好用 by='列名'，比索引更安全
        self.df_sort = self.df.sort_values(by=self.df.columns[1], ascending=False).copy()
        self.logger.info(f"已排序，共 {len(self.df_sort)} 行数据")

        # === 规则定义 (保持不变) ===
        valve_type_rules = {
            "球阀": "Q", "蝶阀": "D", "截止阀": "J", "止回阀": "H",
            "闸阀": "G", "疏水阀": "S", "放料阀": "F", "减压阀": "TP",
            "针阀": "Z", "安全阀": "A" # 注意：去掉重复的减压阀TP，避免冲突
        }
        valve_connect_rules = {"法兰": "0", "卡箍": "1", "焊接": "2", "螺纹": "3", "对夹": "4"}
        valve_seat_rules = {"EPDM": "X", "PTFE": "F", "硬密封": "Y"}
        valve_body_rules = {
            "304": "1E", "316L": "2E", "2205": "3E", "碳钢": "C",
            "搪瓷": "2C", "UPVC": "2G", "TA2": "1T", "C-F": "1C"
        }
        screw_to_DN = {
            "G1/2": "15", "G3/4": "20", "G1-1/4": "32", "G1-1/2": "40"
        }

        # === 确保列存在 ===
        required_cols = ['阀门名称', '密封形式', '阀门规格', '阀门材质', '阀门形式', '阀板材质', '出口管径']
        for col in required_cols:
            if col not in self.df_sort.columns:
                self.logger.warning(f"缺失必要列: {col}，已自动补空值")
                self.df_sort[col] = ""

        sku_list = []
        material_list = []
        # 遍历每一行
        for index, row in self.df_sort.iterrows():
            try:
                # 内部定义取值函数
                def safe_get(col):
                    val = row.get(col, "")
                    return str(val).strip() if pd.notna(val) else ""

                # 获取数据
                valve_name = safe_get('阀门名称')
                valve_spec = safe_get('阀门规格') 
                valve_material1 = safe_get('阀门材质')
                valve_type = safe_get('阀门形式')
                valve_faban1 = safe_get('阀板材质')
                sealing_type = safe_get('密封形式')
                out_diameter = safe_get('出口管径')
                valve_material = valve_material.replace({"C-F","碳钢衬F4"})
                valve_faban = valve_faban1.replace({"C-F","碳钢衬F4"})

                # 逻辑处理
                match = re.search(r'\d+', valve_spec)
                diameter = match.group() if match else "错误"

                if "电动" in valve_type:
                    drive_code = "1"
                elif "气动" in valve_type:
                    drive_code = "2"
                else:
                    drive_code = "0"

                # 查找编码 (找不到则返回 "错误")
                valve_type_code = next((code for key, code in valve_type_rules.items() if key in valve_type), "错误")
                sealing_code = next((code for key, code in valve_seat_rules.items() if key in sealing_type), "错误")
                connect_code = next((code for key, code in valve_connect_rules.items() if key in valve_type), "错误")
                material_code = next((code for key, code in valve_body_rules.items() if key in valve_material), "错误")

                # 生成 SKU
                sku = ""
                if "蝶阀" in valve_type:
                    if "石墨铸铁" in valve_material:
                        sku = f"{valve_type_code}{connect_code}{drive_code}{sealing_code}-{diameter}-1C/{valve_faban}"
                        medium_code = f"(*).阀体：石墨铸铁；\n（*）.阀板：{valve_faban}；\n（*）.密封：{sealing_type}"
                    else:
                        sku = f"{valve_type_code}{connect_code}{drive_code}{sealing_code}-{diameter}-{material_code}"
                        medium_code = f"(*).阀体：{valve_material}；\n（*）.阀板：{valve_faban}；\n（*）.密封：{sealing_type}"
                elif "螺纹" in valve_type:
                    d_screw = screw_to_DN.get(valve_spec, diameter) 
                    sku = f"{valve_type_code}{connect_code}{drive_code}{sealing_code}-{d_screw}-{material_code}"
                    medium_code = f"(*).阀体：{valve_material}；\n（*）.阀芯：{valve_material}；\n（*）.密封：{sealing_type}"
                elif "V形" in valve_type:
                    sku = f"{valve_type_code}V{connect_code}{drive_code}{sealing_code}-{diameter}-{material_code}"
                    medium_code = f"(*).阀体：{valve_material}；\n（*）.阀芯：{valve_material}；\n（*）.密封：{sealing_type}"
                elif "上展" in valve_type or "减压阀" in valve_type:
                    sku = f"{valve_type_code}{connect_code}{drive_code}{sealing_code}-{diameter}/{out_diameter}-{material_code}"
                    medium_code = f"（*）.过流材质：{valve_material}"
                elif "止回阀" in valve_type:
                    sku = f"{valve_type_code}{connect_code}{drive_code}{sealing_code}-{diameter}-{material_code}"
                    medium_code = f"（*）.阀体：{valve_material}；\n（*）.阀瓣：{valve_material}；\n（*）.密封：{sealing_type}"
                else:
                    sku = f"{valve_type_code}{connect_code}{drive_code}{sealing_code}-{diameter}-{material_code}"
                    medium_code = f"(*).阀体：{valve_material}；\n（*）.阀芯：{valve_material}；\n（*）.密封：{sealing_type}"
                
                sku_list.append(sku)
                material_list.append(medium_code)
                self.df_sort['申购单类型'] = '阀门仪表-阀门'
            except Exception as e:
                self.logger.error(f"行 {index} 生成失败: {str(e)}")
                sku_list.append("生成错误")
        
        # 1. 赋值 (此时长度一定匹配)
        self.df_sort['*SKU编号'] = sku_list
        self.df_sort['材质'] = material_list
        self.df_sort['*申购单类型'] = '阀门仪表-阀门'

        if '阀门名称' in self.df_sort.columns:
            self.df_sort = self.df_sort.sort_values(by='名称')
            self.logger.info("数据按 '名称' 排序完成")
        else:
            self.logger.warning("未找到 '名称' 列，跳过最终排序")

        return self.df_sort

    def generate_parameter(self):
        try:
            protocol_path = self.resource_path("dataset\\阀门选型手册.xlsx")
            
            # 实例化时确保传入了 logger
            matcher = ValveProtocolMatcher(protocol_path)
            
            # 执行预处理
            result = matcher.load_and_preprocess()
            
            if result is True:
                self.df_sort = matcher.run_match(self.df_sort)
                self.logger.info("✅ 参数匹配流程全部完成")
            else:
                # 这里的报错通常在类内部已经打印过了，所以这里只是个状态提示
                self.logger.error("❌ 匹配终止：预处理阶段未通过")             
        except Exception as e:
            # ✅ 如果这里报错，说明是传参或者实例化过程就挂了
            self.logger.error(f"❌ 调用匹配器时发生顶层异常: {str(e)}")

        try:
            medium_path = self.resource_path("dataset\\medium.xlsx")
            filler = ParameterFiller(medium_path)
            df_sort = filler.fill_dataframe(self.df_sort)
            self.logger.info("✅ 参数填充流程完成")
        except Exception as e:
            self.logger.error(f"❌ 参数填充时发生异常: {str(e)}")
    
    def merge_by_SKU(self):
        self.df_sort['备注'] = self.df_sort['备注'].fillna('')
        aggregation = {
            '项目号': 'first',
            '阀门形式': 'first',
            '参数': 'first',
            '材质': 'first',
            '申购单备注': 'first',
            '备注': lambda x:'\n'.join(item for item in x if item.strip() != ''),
            '申购人': 'first',
            '申购日期': 'first',
            '*申购单类型': 'first',
            '计数': 'sum',
            '协议号': 'first'
        }
        cols_need = list(aggregation.keys())
        df_clean  = self.df_sort[['*SKU编号']+cols_need].copy()
        self.df_group = df_clean.groupby('*SKU编号',as_index=False).agg(aggregation)
        self.df_group['*申请数量'] = self.df_group['计数']
        self.df_group['所属项目'] = self.df_group['项目号']    
        template_path = self.resource_path("dataset\\申购单导入模板.xlsx")
        template_df =  pd.read_excel(template_path, engine='openpyxl',header = 1)
        template_columns = [str(col).strip().replace('\n', '').replace('\r', '') for col in template_df.columns]
        for col in template_columns:
            if col not in self.df_group.columns:
                # 特殊处理：将 UI 输入的名称映射到模板要求的名称
                if col == "项目编号" and "项目号" in self.df_group.columns:
                    self.df_group["项目编号"] = self.df_group["项目号"]
                elif col == "需求日期" and "申购日期" in self.df_group.columns:
                    self.df_group["需求日期"] = self.df_group["申购日期"]
                elif col == "申请日期" and "申购日期" in self.df_group.columns:
                    self.df_group["申请日期"] = self.df_group["申购日期"]
                elif col=="产品名称" and "阀门形式" in self.df_group.columns:
                    self.df_group["产品名称"] = self.df_group["阀门形式"]
                elif col=="战略合作协议序号" and "协议号" in self.df_group.columns:
                    self.df_group["战略合作协议序号"] = self.df_group["协议号"]
                else:
                    # 💡 这就是你要的效果：模板要求但数据没有的列，统一填空
                    self.df_group[col] = ""
        self.df_output = self.df_group[template_columns].reset_index(drop = True)

        book = load_workbook(template_path)
        sheet = book.active
        start_row = 3
        for row_idx, row in self.df_output.iterrows():
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=start_row + row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                if col_idx == sheet.max_column:  # 如果是最后一列，设置自动换行
                    cell.alignment = Alignment(wrap_text=True)
        self.logger.info(f"✅ 已生成申购单")
        self.book = book
        return book 




                        


        
                    
                
                

                




